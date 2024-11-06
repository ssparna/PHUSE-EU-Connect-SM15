"""
Contains the AnnotationExporter class which contains the logic for exporting annotations
"""
import os
import logging as lg
from sqlite3 import connect, Connection, Cursor
import PyPDF2
import PyPDF2.generic
import openpyxl as pyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill
from openpyxl.cell.cell import Cell
from .generic import PDF, Annotation, Page

class AnnotationExporter:
    """
    Responsible for exporting annotations from a pdf to an \n 
    excel file
    """
    def __init__(self) -> None:
        """
        initializes variables for use in the programm
        """
        self.green_cell_fill = PatternFill(
            start_color = "FF00FF00",
            end_color = "FF00FF00",
            fill_type = "solid")
        self.reset_cell_fill = PatternFill(
            fill_type=None,
            start_color="FFFFFFFF",
            end_color="FF000000")
        self.wb: pyxl.Workbook
        self.pdf: PDF
        self.output_folder: str
        self.exporter_col_ds: str | None
        self.exporter_col_var: str | None
        self.ws_datasets: Worksheet
        self.ws_variables: Worksheet
        self.supp_var_names: list[str] = ["QVAL", "QNAM", "QLABEL"]
        self.ds_replace_annots: list[dict] = []
        self.current_page: Page
        lg.basicConfig(
            filename=f"{os.path.dirname(__file__)}/Annotation_Exporter.log",
            encoding="utf-8",
            level=lg.DEBUG,
            filemode="w")

    def determine_exporter_col(self, sheet: str) -> str | None:
        """
        Determines the exporter column. Returns the exel column index of the free column or None.
        The column chosen is the first free column in the sheet. The exporter column is used for 
        marking the entries as present.

        :param sheet: name of the sheet
        :type sheet: str
        :return: the exel column index of the free column or None
        :rtype: str | None
        """
        value = None
        for cell in self.wb[sheet]["1"]:
            if cell.value is None:
                cell.value = "Present in aCRF"
                value =  "".join([i for i in cell.coordinate if not i.isdigit()])  # remove all digits

        if value is None:
            lg.critical("no free column for sheet %s found, exiting...", sheet)
            exit()

        return value

    def export_annotations(self, template_path: str, pdf_path: str, output_folder: str) -> None:
        """
        Exports annots, this is the main function that should be called. 
        Expects the paths to have the correct endings (.pdf, .xlsx).

        :param template_path: path to the template file
        :type template_path: str
        :param pdf_path: path to the pdf file
        :type pdf_path: str
        :param output_folder: path to the output folder
        :type output_folder: str
        """
        print("exporting annotations...")
        lg.info("export annots")
        self.wb = pyxl.load_workbook(template_path)
        self.output_folder = output_folder

        self.exporter_col_ds = self.determine_exporter_col("Datasets")
        self.exporter_col_var = self.determine_exporter_col("Variables")

        self.ws_datasets = self.wb["Datasets"]
        self.ws_variables = self.wb["Variables"]

        self.pdf: PDF = PDF(PyPDF2.PdfReader(pdf_path))

        for page in self.pdf.pages:
            self.current_page = page
            lg.info("starting on page: %s", page.get_page_nr())

            self.add_to_workbook(page.get_annotations())

            lg.debug(page.get_datasets())
            lg.info("Page %s done!", page.get_page_nr())


        self.wb.save(f"{output_folder}/output.xlsx")
        print("generating csv...")
        lg.info("generating csv of export")
        self.generate_variable_csv()
        self.generate_dataset_csv()

        print("complete!")
        lg.info("exported annots")

    def generate_sqlite(self, output_folder: str) -> None:
        """
        generates an sqlite database from the annotations and
        saves it in the output folder

        :param output_folder: path to the output folder
        :type output_folder: str
        """
        conn: Connection= connect(f"{output_folder}/annotations.sqlite")
        c: Cursor = conn.cursor()
        c.execute("DROP TABLE IF EXISTS annotations")
        c.execute("""CREATE TABLE annotations
                    (id INTEGER PRIMARY KEY AUTOINCREMENT,
                    dataset BOOLEAN,
                    new_dataset BOOLEAN,
                    dataset_name TEXT,
                    supp BOOLEAN,
                    assigned_dataset TEXT,
                    variable_name TEXT,
                    content TEXT,
                    color TEXT,
                    page_number INTEGER)""")

        for page in self.pdf.pages:
            for annot in page.get_annotations():
                if annot.is_valid:
                    c.execute("""INSERT INTO annotations
                        (dataset, new_dataset, dataset_name, supp, assigned_dataset, variable_name, content, color, page_number)
                        VALUES (?,?,?,?,?,?,?,?,?)""",
                        (annot.dataset,
                         annot.new_datset,
                         annot.dataset_name,
                         annot.supp,
                         annot.assigned_dataset,
                         annot.variable_name,
                         annot.content,
                         str(annot.color),
                         annot.page.get_page_nr() + 1))

        conn.commit()

    def enter_dataset(self, annot: Annotation) -> None:
        """
        Adds a dataset to the Datasets shhet in the workbook

        :param annot: annotation object
        :type annot: Annotation
        """
        for cell in self.ws_datasets.iter_rows(max_col=1):
            cell = cell[0]
            if cell.value == annot.dataset_name:
                y_coordinate = cell.coordinate.split("A", 1)[1]
                self.ws_datasets[f"{self.exporter_col_ds}{y_coordinate}"] = "Present"
                self.ws_datasets[f"{self.exporter_col_ds}{y_coordinate}"].fill = self.green_cell_fill

                lg.debug(
                    "%s was assigned as a dataset with the color %s",
                    annot.dataset_name, annot.color)
                return

        self.ws_datasets.append({
            "A": annot.dataset_name,
            self.exporter_col_ds: "Present",
        })
        self.ws_datasets[self.exporter_col_ds][self.ws_datasets.max_row - 1].fill = self.green_cell_fill
        self.ws_datasets["A"][self.ws_datasets.max_row - 1].fill = self.reset_cell_fill

    def enter_supp(self, annot: Annotation) -> None:
        """
        adds a SUPPxx dataset to datasets and the three SUPPxx variables to variables

        :param annot: annotation object
        :type annot: Annotation
        """
        self.enter_dataset(annot)

        modified_cols: list[str] = ["B", "C", "L", "F"]

        all_values = [x[0] for x in self.ws_variables.iter_rows(min_col=2, max_col=2, values_only=True)] # very performance inefficient

        if annot.dataset_name in all_values:
            for cell in self.ws_variables["B"]:
                if cell.value != annot.dataset_name:
                    continue
                self.add_page_cell(self.ws_variables[f"M{cell.row}"])
                break
        else:
            for var_name in self.supp_var_names:
                self.ws_variables.append({
                    "B": annot.dataset_name,
                    "C": var_name,
                    "L": "CRF",
                    "F": "200",
                    self.exporter_col_var: "Present"})

                self.add_page_cell(self.ws_variables[f"M{self.ws_variables.max_row}"])
                self.ws_variables[self.exporter_col_var][self.ws_variables.max_row - 1].fill = self.green_cell_fill

                for col in modified_cols:
                    self.ws_variables[col][self.ws_variables.max_row - 1].fill = self.reset_cell_fill

    def add_to_workbook(self, annotations: list[Annotation]) -> None:
        """
        adds both datasets and variables to the workbook.
        Takes a list of annotations, desigend to work with PDF.pages
        
        :param data: list of annotations
        :type data: list[Annotation]
        """
        for annot in annotations:
            if annot.dataset:
                self.enter_dataset(annot)
            elif annot.supp:
                self.enter_supp(annot)
            else:
                self.enter_variable(annot)

    def enter_variable(self, annot: Annotation) -> None:
        """
        adds a variable to the workbook

        :param annot: annotation object
        :type annot: Annotation
        """
        annot.sort_into_datasets()
        if annot.assigned_dataset is None:
            return

        for cell in self.ws_variables["B"]:
            y_coordinate = cell.coordinate.split("B", 1)[1]

            if cell.value == annot.assigned_dataset and self.ws_variables["C" + y_coordinate].value == annot.variable_name:
                if self.ws_variables[f"{self.exporter_col_var}{y_coordinate}"].value == "Present":
                    self.add_page_cell(self.ws_variables[f"M{y_coordinate}"])
                    return

                self.ws_variables[f"{self.exporter_col_var}{y_coordinate}"].value = "Present"
                self.ws_variables[f"{self.exporter_col_var}{y_coordinate}"].fill = self.green_cell_fill
                self.ws_variables[f"L{y_coordinate}"].value = "CRF"
                self.add_page_cell(self.ws_variables[f"M{y_coordinate}"])
                return

        self.ws_variables.append({
            "B": annot.assigned_dataset,
            "C": annot.variable_name,
            "L": "CRF",
            "F": "200",
            self.exporter_col_var: "Present"
        })
        self.add_page_cell(self.ws_variables[f"M{self.ws_variables.max_row}"])
        self.ws_variables[self.exporter_col_var][self.ws_variables.max_row - 1].fill = self.green_cell_fill
        for modified_col in ["B", "C", "L", "F"]:
            self.ws_variables[modified_col][self.ws_variables.max_row - 1].fill = self.reset_cell_fill

    def generate_variable_csv(self) -> None:
        """
        generates the csv for the variables and saves it in the output folder
        """
        csv_list = ["Variable Name#Variable Label#Dataset Name#Page(s)\n"] # start with first line
        for cell in self.ws_variables[self.exporter_col_var]:
            if cell.value == "Present":
                cell_str: str = ""
                cell_str += str(self.ws_variables["C" + str(cell.row)].value) + "#"
                cell_str += str(self.ws_variables["D" + str(cell.row)].value) + "#"
                cell_str += str(self.ws_variables["B" + str(cell.row)].value) + "#"
                cell_str += str(self.ws_variables["M" + str(cell.row)].value) + "\n"
                csv_list.append(cell_str)


        csv_str = "".join(csv_list)
        with open(f"{self.output_folder}/Variables.csv", "w", encoding="utf-8") as f:
            f.write(csv_str)

    def generate_dataset_csv(self):
        """
        generates the csv for the datasets and saves it in the output folder
        """
        csv_list: list[str] = ["Dataset Name#Color\n"] # start with first line
        for page in self.pdf.pages:
            for dataset in page.get_datasets():
                csv_entry: str = f"{dataset[0]}#{dataset[1]}\n"
                if csv_entry in csv_list:
                    continue
                csv_list.append(csv_entry)

        csv_str = "".join(csv_list)

        with open(f"{self.output_folder}/Datasets.csv", "w", encoding="utf-8") as f:
            f.write(csv_str)

    def add_page_cell(self, cell: Cell):
        """
        Appends current page number to the string in a cell if it is not already present.

        :param cell: The cell in which to add the current page number
        :type cell: Cell
        """
        if not cell.value:
            new_value = str(self.current_page.get_page_nr() + 1)
            cell.value = new_value
            cell.fill = self.reset_cell_fill
        elif str(self.current_page.get_page_nr() + 1) in cell.value:
            return
        else:
            new_value = f"{cell.value} {self.current_page.get_page_nr() + 1}"
